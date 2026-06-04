#!/usr/bin/env python3
"""Audit availability of a deterministic WC2026 knockout bracket mapping."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.simulation.knockout_bracket import ensure_manual_bracket_template

ROOT = Path(__file__).parent.parent
WORKBOOK = ROOT / "Rules of the game" / "FIF8A World Cup 2026_Player_template.xlsx"
GUIDE = ROOT / "Rules of the game" / "FIF8A World Cup 2026 - Player guide.pdf"
MAPPING = ROOT / "data" / "reference" / "knockout_bracket_mapping.csv"
MANUAL_TEMPLATE = ROOT / "data" / "reference" / "knockout_bracket_mapping_manual.csv"
REPORT = ROOT / "outputs" / "reports" / "bracket_structure_audit.md"


def workbook_evidence() -> list[str]:
    if not WORKBOOK.exists():
        return ["Workbook missing."]
    workbook = load_workbook(WORKBOOK, data_only=False, read_only=True)
    evidence = [f"Workbook sheets: {', '.join(workbook.sheetnames)}."]
    if "Last 8 teams predictions" in workbook.sheetnames:
        sheet = workbook["Last 8 teams predictions"]
        labels = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in {"Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7", "Q8", "S1", "S2", "S3", "S4", "F1", "F2", "W"}:
                    labels.append(str(cell.value))
        evidence.append(f"`Last 8 teams predictions` contains stage labels only: {', '.join(labels)}.")
    if "Best 3rds calculation" in workbook.sheetnames:
        sheet = workbook["Best 3rds calculation"]
        source_positions = [sheet[f"C{row}"].value for row in range(15, 27)]
        evidence.append(
            "`Best 3rds calculation` ranks third-place teams from "
            f"{', '.join(str(value) for value in source_positions)} but does not map them to R32 slots."
        )
    return evidence


def main() -> None:
    ensure_manual_bracket_template(MANUAL_TEMPLATE)
    exact_mapping_available = MAPPING.exists()

    lines = [
        "# Bracket Structure Audit",
        "",
        f"- Exact Round-of-32 mapping available: **{exact_mapping_available}**",
        f"- Expected explicit mapping file: `{MAPPING.relative_to(ROOT)}`",
        f"- Manual template created/available: `{MANUAL_TEMPLATE.relative_to(ROOT)}`",
        f"- Rules workbook inspected: `{WORKBOOK.relative_to(ROOT)}` exists = **{WORKBOOK.exists()}**",
        f"- Player guide inspected: `{GUIDE.relative_to(ROOT)}` exists = **{GUIDE.exists()}**",
        "",
        "## Findings",
        "",
        "- The rules describe a 48-team tournament with top two from each group plus the eight best third-placed teams advancing to a Round of 32.",
        "- The workbook includes a `Best 3rds calculation` tab that ranks third-placed teams deterministically by group-table metrics.",
        "- The workbook `Last 8 teams predictions` tab contains only Q1-Q8/S1-S4/F1-F2/W entry slots, not source-group-to-Round-of-32 bracket assignments.",
        "- No local reference/manual/FIFA cached file exposes how A1/B2/best-third slots are placed into the Round of 32 bracket.",
        "",
        "## Evidence",
        "",
    ]
    lines.extend(f"- {item}" for item in workbook_evidence())
    lines.extend(
        [
            "",
            "## Consequence",
            "",
            "- Exact path-aware full-tournament simulation is blocked until an official or manually reviewed bracket mapping is provided.",
            "- The mapping for best third-placed teams is especially important: best-thirds can be selected deterministically, but their R32 placement is not available locally.",
            "- Manual bracket input is required in `data/reference/knockout_bracket_mapping_manual.csv` before producing path-aware Last-8 probabilities.",
            "",
            "## Required Manual Mapping",
            "",
            "Populate each R32 row with a `source_group_position` such as `A1`, `B2`, or a reviewed best-third slot convention, then copy/rename the reviewed file to `data/reference/knockout_bracket_mapping.csv`.",
            "Do not infer these slots from team strength or expected rankings.",
        ]
    )
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {REPORT.relative_to(ROOT)}")
    print(f"Wrote {MANUAL_TEMPLATE.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
